import os
import sys
import json
import argparse
import traceback
import pandas as pd

# Try to find build_pdf_template no matter where you place this helper.
# 1) Prefer a function already defined in the same module (globals()).
# 2) Else try to import it from pdf_prefill (common case in your project).
_build_pdf_template_external = None
_build_pdf_template_import_error = None
_build_pdf_template_import_tb = None
try:
    from pdf_prefill import build_pdf_template as _build_pdf_template_external  # type: ignore
except Exception as _e:
    _build_pdf_template_external = None
    _build_pdf_template_import_error = _e
    _build_pdf_template_import_tb = traceback.format_exc()


def _get_builder():
    """Return a callable build_pdf_template or raise a helpful error with root cause."""
    fn = globals().get("build_pdf_template", None)
    if callable(fn):
        return fn
    if callable(_build_pdf_template_external):
        return _build_pdf_template_external

    detail = ""
    if _build_pdf_template_import_error is not None:
        # Print full traceback to stderr so the real cause (syntax/indent) is visible immediately.
        sys.stderr.write("----- pdf_prefill import traceback -----\n")
        if _build_pdf_template_import_tb:
            sys.stderr.write(_build_pdf_template_import_tb + "\n")
        else:
            sys.stderr.write(repr(_build_pdf_template_import_error) + "\n")
        sys.stderr.write("----- end traceback -----\n")
        detail = (
            f"\n(Import error: {type(_build_pdf_template_import_error).__name__}: "
            f"{_build_pdf_template_import_error})"
        )

    raise RuntimeError(
        "build_pdf_template(...) was not found. "
        "Define it above this code OR ensure `from pdf_prefill import build_pdf_template` works."
        + detail
    )


def _ensure_parent_dir(path: str):
    d = os.path.dirname(os.path.abspath(path))
    if d and not os.path.exists(d):
        os.makedirs(d, exist_ok=True)


def _field_title_from_fdef(fdef: dict) -> str:
    """Prefer a concise key when available, otherwise fallback to label."""
    # For checkbox bullets we created `label_short` like 'For All Subscribers – item 1'
    title = (fdef.get("label_short") or fdef.get("label") or "").strip()

    # If extremely long, keep the text before the first colon (common pattern).
    if ":" in title:
        left = title.split(":", 1)[0].strip()
        # Only shorten when the left side is meaningful
        if len(left) >= 10:
            title = left

    return title


def _choices_str_from_fdef(fdef: dict) -> str:
    """
    Return a semicolon-separated string of choices for acro_choice fields,
    or empty string otherwise. Filters Nones and trims whitespace.
    """
    placement = (fdef.get("placement") or "").lower()
    choices = fdef.get("choices")
    if placement == "acro_choice" and isinstance(choices, list) and choices:
        return "; ".join(str(x).strip() for x in choices if x is not None)
    return ""


def _try_add_excel_dropdowns(xlsx_path: str):
    """
    Add per-row data validation lists on the Value column for rows that have a
    short 'Choices' list. This is best-effort and quietly does nothing if
    openpyxl isn't available or anything goes wrong.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.worksheet.datavalidation import DataValidation
    except Exception:
        # openpyxl not installed; skip silently
        return

    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active  # single-sheet export
        # Map headers -> column indices
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        col_value = headers.get("Value")
        col_choices = headers.get("Choices")
        if not col_value or not col_choices:
            wb.close()
            return

        last_row = ws.max_row

        for r in range(2, last_row + 1):
            ch_raw = ws.cell(row=r, column=col_choices).value
            if not ch_raw or not str(ch_raw).strip():
                continue

            # Build a CSV list for Excel data validation (max 255 chars)
            items = [s.strip() for s in str(ch_raw).split(";") if s.strip()]
            # Escape double-quotes by doubling them (Excel formula string rule)
            safe_items = [it.replace('"', '""') for it in items]
            csv_list = ",".join(safe_items)

            if not csv_list or len(csv_list) > 255:
                # Too long for inline list validation; skip this row
                continue

            dv = DataValidation(type="list", formula1=f'"{csv_list}"', allow_blank=True, showDropDown=True)
            addr = ws.cell(row=r, column=col_value).coordinate
            dv.add(ws[addr])
            ws.add_data_validation(dv)

        wb.save(xlsx_path)
        wb.close()
    except Exception:
        # best-effort; ignore errors so main flow isn't interrupted
        pass


def export_lookup_template_from_json(template_json: str,
                                     out_path: str = "lookup_template.xlsx") -> str:
    """
    Produce an Excel (or CSV if path ends with .csv) with columns:
    Section | Page | Field | Index | Value | Choices

    Returns the output path.
    """
    if not os.path.exists(template_json):
        raise FileNotFoundError(f"Template JSON not found: {template_json}")

    with open(template_json, "r", encoding="utf-8") as f:
        tpl = json.load(f)

    rows = []
    for fdef in tpl.get("fields", []) or []:
        # Skip junk placeholders
        raw_label = (fdef.get("label") or "").strip()
        if not raw_label or raw_label.startswith("unknown_"):
            continue

        section = (fdef.get("section") or "").strip()
        try:
            page = int(fdef.get("page", 0) or 0)
        except Exception:
            page = 0
        try:
            index = int(fdef.get("index", 1) or 1)
        except Exception:
            index = 1

        field = _field_title_from_fdef(fdef)
        choices_str = _choices_str_from_fdef(fdef)  # <-- NEW

        rows.append({
            "Section": section,
            "Page": page,
            "Field": field,
            "Index": index,
            "Value": "",              # you fill this in later
            "Choices": choices_str,   # <-- NEW COLUMN
        })

    # Stable ordering for human-friendly editing
    rows.sort(key=lambda r: (r["Page"], r["Section"].lower(), r["Field"].lower(), r["Index"]))

    df = pd.DataFrame(rows, columns=["Section", "Page", "Field", "Index", "Value", "Choices"])

    _ensure_parent_dir(out_path)

    # Write Excel or CSV (fallback to CSV if openpyxl not installed)
    if out_path.lower().endswith(".csv"):
        df.to_csv(out_path, index=False, encoding="utf-8-sig", newline="")
    else:
        try:
            # requires openpyxl for .xlsx
            df.to_excel(out_path, index=False)
            # Best-effort: add Excel dropdowns to Value based on Choices (short lists only)
            _try_add_excel_dropdowns(out_path)
        except Exception as e:
            # fallback
            fallback = os.path.splitext(out_path)[0] + ".csv"
            df.to_csv(fallback, index=False, encoding="utf-8-sig", newline="")
            print(f"⚠️  Could not write Excel ({e}). Wrote CSV instead: {fallback}")
            return fallback

    print(f"📤 Lookup template written to {out_path} with {len(df)} rows.")
    return out_path


def _call_builder_with_compat(builder, input_pdf: str, template_json: str):
    """
    Call build_pdf_template with broad signature compatibility:
    - build_pdf_template(path, template_json, lookup_rows=None, dry_run=False)
    - build_pdf_template(path, template_json=..., lookup_rows=None, dry_run=False)
    - build_pdf_template(doc_or_path=path, template_json=..., ...)
    """
    _ensure_parent_dir(template_json)

    # Try the most common positional signature first
    try:
        return builder(input_pdf, template_json, lookup_rows=None, dry_run=False)
    except TypeError:
        pass

    # Try keyword style with template_json=
    try:
        return builder(input_pdf, template_json=template_json, lookup_rows=None, dry_run=False)
    except TypeError:
        pass

    # Try fully keyworded with doc_or_path=
    try:
        return builder(doc_or_path=input_pdf, template_json=template_json, lookup_rows=None, dry_run=False)
    except TypeError:
        pass

    # Last attempt: just pass (path, template_json) and let the callee ignore extras
    try:
        return builder(input_pdf, template_json)
    except Exception as e:
        raise RuntimeError(
            "Found build_pdf_template but could not call it with a compatible signature. "
            "Please ensure it accepts either (path, template_json, ...) or (doc_or_path=..., template_json=...)."
        ) from e


def export_lookup_template_from_pdf(input_pdf: str,
                                    template_json: str = "template_fields.json",
                                    out_path: str = "lookup_template.xlsx",
                                    rebuild_template: bool = False) -> str:
    """
    Ensures a template JSON exists (building it if needed), then exports the lookup sheet.
    """
    must_build = rebuild_template or not os.path.exists(template_json)
    if must_build:
        builder = _get_builder()
        _call_builder_with_compat(builder, input_pdf, template_json)

    return export_lookup_template_from_json(template_json, out_path)


# ---------------------------
# CLI
# ---------------------------
def main():
    ap = argparse.ArgumentParser(
        description="Export a blank lookup sheet (Section, Page, Field, Index, Value, Choices) from template_fields.json"
    )
    ap.add_argument("--input", required=True, help="Input PDF path")
    ap.add_argument("--template", default="template_fields.json", help="Template JSON (built if missing)")
    ap.add_argument("--make-lookup", metavar="OUT.xlsx",
                    help="Path to write the blank lookup sheet (xlsx or csv).")
    ap.add_argument("--rebuild-template", action="store_true",
                    help="Force rebuild of the template JSON even if it already exists")
    args = ap.parse_args()

    if args.make_lookup:
        export_lookup_template_from_pdf(
            input_pdf=args.input,
            template_json=args.template,
            out_path=args.make_lookup,
            rebuild_template=args.rebuild_template,
        )
        return

    # If the user didn’t pass --make-lookup, just print a short hint
    print("Nothing to do. Pass --make-lookup OUT.xlsx (or .csv) to export a blank lookup sheet.")


if __name__ == "__main__":
    main()